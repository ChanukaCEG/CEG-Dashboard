import streamlit as st
import pandas as pd
import json
import io
import base64 as b64lib
from datetime import datetime
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import tempfile
import os
import requests
import re

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                 Paragraph, Spacer, Image as RLImage)
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

st.set_page_config(page_title="CEG Vendor Purchase Analysis", layout="wide")

import os as _os
_LOGO_B64 = ""
_logo_src = "ceg_logo.png"
if _os.path.exists(_logo_src):
    with open(_logo_src, "rb") as _lf:
        _LOGO_B64 = b64lib.b64encode(_lf.read()).decode()

COL = {
    "t1":"1F3864","t2":"005F73","t3":"4A235A",
    "t4":"784212","t5":"1A5276",
    "r1a":"D9E1F2","r1b":"EBF0FA",
    "r2a":"D4EFEA","r2b":"E8F8F5",
    "r3a":"E8DAEF","r3b":"F5EEF8",
    "r4a":"FAD7A0","r4b":"FDEBD0",
    "r5a":"D6EAF8","r5b":"EBF5FB",
    "yrt":"AED6F1","tot":"C0C0C0",
}

GDRIVE_FILE_ID = "11kqP7ybyCupBMjSTdFsiP2LpoSJ6KKJY"

@st.cache_data
def load_data():
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    session = requests.Session()
    url = f"https://drive.google.com/uc?export=download&id={GDRIVE_FILE_ID}"
    response = session.get(url, stream=True)
    token = None
    for key, value in response.cookies.items():
        if key.startswith("download_warning"):
            token = value
    if not token:
        for key, value in response.cookies.items():
            if "warning" in key.lower():
                token = value
    if not token:
        content = response.content.decode("utf-8", errors="ignore")
        match = re.search(r'confirm=([0-9A-Za-z_]+)', content)
        if match:
            token = match.group(1)
    if token:
        url = f"https://drive.google.com/uc?export=download&confirm={token}&id={GDRIVE_FILE_ID}"
    else:
        url = f"https://drive.google.com/uc?export=download&confirm=t&id={GDRIVE_FILE_ID}"
    response = session.get(url, stream=True)
    with open(tmp.name, "wb") as f:
        for chunk in response.iter_content(32768):
            if chunk:
                f.write(chunk)
    ile = pd.read_excel(tmp.name, sheet_name="ILE DATA")
    ile = ile[
        (ile["Entry_Type"] == "Purchase") &
        (ile["Document_Type"].isin(["Purchase Receipt","Purchase Return Shipment"]))
    ]
    ile["eff_cost"] = ile["Cost_Amount_Actual"].fillna(0)
    mask = ile["eff_cost"] == 0
    ile.loc[mask, "eff_cost"] = ile.loc[mask, "Cost_Amount_Expected"].fillna(0)
    ile["row_unit_cost"] = ile.apply(
        lambda r: r["eff_cost"] / r["Quantity"]
        if r["Quantity"] and r["Quantity"] != 0 else 0, axis=1)
    ile["Posting_Date"] = pd.to_datetime(ile["Posting_Date"])
    ile["Year"]         = ile["Posting_Date"].dt.year
    ile["YearMonth"]    = ile["Posting_Date"].dt.to_period("M").astype(str)
    ile["Month"]        = ile["Posting_Date"].dt.month
    vendors = pd.read_excel(tmp.name, sheet_name="VENDOR CARD")
    vendors = vendors[["No","Name"]].rename(
        columns={"No":"Vendor_Code","Name":"Vendor_Name"})
    ile = ile.merge(vendors, left_on="Source_No",
                    right_on="Vendor_Code", how="left")
    os.unlink(tmp.name)
    return ile, vendors

ile_df, vendor_df = load_data()

all_vendor_names = ["ALL"] + sorted(vendor_df["Vendor_Name"].dropna().unique().tolist())
all_vendor_codes = ["ALL"] + sorted(vendor_df["Vendor_Code"].dropna().unique().tolist())
all_years = ["ALL"] + sorted(ile_df["Year"].dropna().unique().astype(str).tolist())

for key,val in {
    "vname_select":"ALL","vcode_select":"ALL","year_select":"ALL",
    "f_t1_n":10,"f_t1_all":False,
    "f_t2_n":10,"f_t2_all":False,
    "f_t3_n":10,"f_t3_all":False,
    "f_t4_srch":"","f_t4_abc":"ALL",
    "xl_ready":False,"pdf_ready":False,
    "xl_data":None,"pdf_data":None,
    "xl_fname":"","pdf_fname":"",
}.items():
    if key not in st.session_state:
        st.session_state[key] = val

def on_name_change():
    name = st.session_state.vname_select
    if name != "ALL":
        matched = vendor_df[vendor_df["Vendor_Name"] == name]["Vendor_Code"].values
        st.session_state["vcode_select"] = matched[0] if len(matched) > 0 else "ALL"
    else:
        st.session_state["vcode_select"] = "ALL"

def on_code_change():
    code = st.session_state.vcode_select
    if code != "ALL":
        matched = vendor_df[vendor_df["Vendor_Code"] == code]["Vendor_Name"].values
        st.session_state["vname_select"] = matched[0] if len(matched) > 0 else "ALL"
    else:
        st.session_state["vname_select"] = "ALL"

if _LOGO_B64:
    logo_html = f'<img src="data:image/png;base64,{_LOGO_B64}" style="height:45px;vertical-align:middle;margin-right:15px;">'
else:
    logo_html = "&#9729; "

st.markdown(f"""
    <div style='background-color:#1F3864;padding:12px 20px;border-radius:5px;
                margin-bottom:10px;display:flex;align-items:center;justify-content:center'>
        {logo_html}
        <h2 style='color:white;margin:0;font-size:38px'>
            Cloud Energy Gas (CEG) &mdash; Vendor Purchase Analysis
        </h2>
    </div>
""", unsafe_allow_html=True)

cv1,cv2,cv3,cv4,cv5,cv6 = st.columns([1.2,1.2,0.7,0.9,0.9,0.6])
with cv1:
    st.markdown("<p style='font-size:12px;font-weight:bold;margin-bottom:2px'>Vendor Name</p>", unsafe_allow_html=True)
    st.selectbox("",all_vendor_names,key="vname_select",
                  on_change=on_name_change,label_visibility="collapsed")
with cv2:
    st.markdown("<p style='font-size:12px;font-weight:bold;margin-bottom:2px'>Vendor Code</p>", unsafe_allow_html=True)
    st.selectbox("",all_vendor_codes,key="vcode_select",
                  on_change=on_code_change,label_visibility="collapsed")
with cv3:
    st.markdown("<p style='font-size:12px;font-weight:bold;margin-bottom:2px'>Year</p>", unsafe_allow_html=True)
    st.selectbox("",all_years,key="year_select",label_visibility="collapsed")
with cv4:
    st.markdown("<p style='font-size:12px;font-weight:bold;margin-bottom:2px'>Export Tables</p>", unsafe_allow_html=True)
    export_tables = st.multiselect("",
        ["Table 1 - Top X Qty","Table 2 - Freq Purchased",
         "Table 3 - Price Analysis","Table 4 - ABC Analysis",
         "Table 5 - Monthly Trend","Charts - Year on Year"],
        default=[],label_visibility="collapsed")
with cv5:
    st.markdown("<p style='font-size:12px;font-weight:bold;margin-bottom:2px'>Export Format</p>", unsafe_allow_html=True)
    export_fmt = st.multiselect("",["Excel","PDF"],
        default=[],label_visibility="collapsed")
with cv6:
    st.markdown("<p style='font-size:12px;font-weight:bold;margin-bottom:2px'>&#8203;</p>", unsafe_allow_html=True)
    do_export = st.button("⬇ Export", use_container_width=True)

filtered_df = ile_df.copy()
if st.session_state.vcode_select != "ALL":
    filtered_df = filtered_df[
        filtered_df["Source_No"] == st.session_state.vcode_select]
if st.session_state.year_select != "ALL":
    filtered_df = filtered_df[
        filtered_df["Year"] == int(st.session_state.year_select)]

def calc_tables(df):
    t1 = (df.groupby(["Item_No","Item_Description"])
          .agg(TQty=("Quantity","sum"))
          .reset_index()
          .rename(columns={"Item_No":"ItemNo","Item_Description":"Desc"})
          .sort_values("TQty",ascending=False).reset_index(drop=True))
    t1["Rank"] = t1.index+1
    t1["TQty"] = t1["TQty"].round(0).astype(int)

    t2 = (df.groupby(["Item_No","Item_Description"])
          .agg(TQty=("Quantity","sum"),NPOs=("Document_No","nunique"))
          .reset_index()
          .rename(columns={"Item_No":"ItemNo","Item_Description":"Desc"}))
    t2["PF"] = (t2["NPOs"]/t2["TQty"]).round(4)
    t2 = t2.sort_values("NPOs",ascending=False).reset_index(drop=True)
    t2["Rank"] = t2.index+1
    t2["TQty"] = t2["TQty"].round(0).astype(int)

    t3 = (df.groupby(["Item_No","Item_Description"])
          .agg(TQty=("Quantity","sum"),TCost=("eff_cost","sum"),
               NPOs=("Document_No","nunique"),
               MinP=("row_unit_cost","min"),MaxP=("row_unit_cost","max"))
          .reset_index()
          .rename(columns={"Item_No":"ItemNo","Item_Description":"Desc"}))
    last_p = (df.sort_values("Posting_Date")
              .groupby(["Item_No","Item_Description"])
              .apply(lambda g: g.loc[g["Posting_Date"].idxmax(),"row_unit_cost"])
              .reset_index()
              .rename(columns={"Item_No":"ItemNo","Item_Description":"Desc",0:"LastP"}))
    t3 = t3.merge(last_p, on=["ItemNo","Desc"], how="left")
    t3["AvgP"] = (t3["TCost"]/t3["TQty"]).round(2)
    t3["PVar"] = (t3["MaxP"]-t3["MinP"]).round(2)
    t3["TCost"] = t3["TCost"].round(0).astype(int)
    t3["TQty"]  = t3["TQty"].round(0).astype(int)
    t3["MinP"]  = t3["MinP"].round(2)
    t3["MaxP"]  = t3["MaxP"].round(2)
    t3["LastP"] = t3["LastP"].round(2)
    t3 = t3.sort_values("TCost",ascending=False).reset_index(drop=True)

    t4 = (df.groupby(["Item_No","Item_Description"])
          .agg(TCost=("eff_cost","sum"))
          .reset_index()
          .rename(columns={"Item_No":"ItemNo","Item_Description":"Desc"})
          .sort_values("TCost",ascending=False).reset_index(drop=True))
    ts = t4["TCost"].sum()
    t4["PctS"] = (t4["TCost"]/ts*100).round(2) if ts!=0 else 0
    t4["CumP"] = t4["PctS"].cumsum().round(2)
    t4["ABC"]  = t4["CumP"].apply(lambda x:"A" if x<=70 else("B" if x<=90 else "C"))
    t4["TCost"]= t4["TCost"].round(0).astype(int)

    t5 = (df.groupby("YearMonth")
          .agg(NPOs=("Document_No","nunique"),TQty=("Quantity","sum"),
               TCost=("eff_cost","sum"))
          .reset_index().sort_values("YearMonth").reset_index(drop=True))
    t5["AvgP"] = (t5["TCost"]/t5["TQty"]).round(2)
    t5["TQty"] = t5["TQty"].round(0).astype(int)
    t5["TCost"]= t5["TCost"].round(0).astype(int)
    t5["Year"] = t5["YearMonth"].str[:4]
    return (t1,t2,t3,t4,t5)

def calc_charts(df):
    ch = (df.groupby(["Year","Month"])
          .agg(TCost=("eff_cost","sum"),TQty=("Quantity","sum"),
               NPOs=("Document_No","nunique"))
          .reset_index())
    ch["TCost"] = ch["TCost"].round(0).astype(int)
    ch["TQty"]  = ch["TQty"].round(0).astype(int)
    years = sorted(ch["Year"].unique().tolist())
    months = list(range(1,13))
    month_names = ["Jan","Feb","Mar","Apr","May","Jun",
                   "Jul","Aug","Sep","Oct","Nov","Dec"]
    result = {"years": years, "months": month_names, "series": {}}
    for yr in years:
        yr_data = ch[ch["Year"]==yr].set_index("Month")
        result["series"][str(yr)] = {
            "cost": [int(yr_data.loc[m,"TCost"]) if m in yr_data.index else 0 for m in months],
            "qty":  [int(yr_data.loc[m,"TQty"])  if m in yr_data.index else 0 for m in months],
            "npos": [int(yr_data.loc[m,"NPOs"])  if m in yr_data.index else 0 for m in months],
        }
    return result

t1,t2,t3,t4,t5 = calc_tables(filtered_df)
chart_data = calc_charts(filtered_df)

def get_export_data():
    e1 = t1 if st.session_state.f_t1_all else t1.head(st.session_state.f_t1_n)
    tot1 = {"TQty": int(e1["TQty"].sum())}
    e2 = t2 if st.session_state.f_t2_all else t2.head(st.session_state.f_t2_n)
    tot2 = {"TQty": int(e2["TQty"].sum()), "NPOs": int(e2["NPOs"].sum())}
    e3 = t3 if st.session_state.f_t3_all else t3.head(st.session_state.f_t3_n)
    tot3 = {"TQty": int(e3["TQty"].sum()), "TCost": int(e3["TCost"].sum())}
    srch = st.session_state.f_t4_srch.lower()
    abc  = st.session_state.f_t4_abc
    e4 = t4.copy()
    if srch:
        e4 = e4[e4["ItemNo"].str.lower().str.contains(srch) |
                e4["Desc"].str.lower().str.contains(srch)]
    if abc == "AB":
        e4 = e4[e4["ABC"].isin(["A","B"])]
    elif abc in ["A","B","C"]:
        e4 = e4[e4["ABC"] == abc]
    tot4 = {"TCost": int(e4["TCost"].sum())}
    e5 = t5.copy()
    tot5 = {"NPOs": int(e5["NPOs"].sum()), "TQty": int(e5["TQty"].sum()),
            "TCost": int(e5["TCost"].sum())}
    return e1,e2,e3,e4,e5,tot1,tot2,tot3,tot4,tot5

tot1_all = {"TQty": int(t1["TQty"].sum())}
tot2_all = {"TQty": int(t2["TQty"].sum()), "NPOs": int(t2["NPOs"].sum())}
tot3_all = {"TQty": int(t3["TQty"].sum()), "TCost": int(t3["TCost"].sum())}
tot4_all = {"TCost": int(t4["TCost"].sum())}
tot5_all = {"NPOs": int(t5["NPOs"].sum()), "TQty": int(t5["TQty"].sum()),
            "TCost": int(t5["TCost"].sum())}

data_json = json.dumps({
    "t1":t1.fillna(0).to_dict("records"),
    "t2":t2.fillna(0).to_dict("records"),
    "t3":t3.fillna(0).to_dict("records"),
    "t4":t4.fillna(0).to_dict("records"),
    "t5":t5.fillna(0).to_dict("records"),
    "tot1":tot1_all,"tot2":tot2_all,"tot3":tot3_all,
    "tot4":tot4_all,"tot5":tot5_all,
    "charts": chart_data,
}, default=str)

COST_COLORS_MPL = ['#1F3864','#85B7EB','#4A235A','#005F73','#1A5276']
QTY_COLORS_MPL  = ['#1B5E20','#66BB6A','#2E7D32','#A5D6A7','#388E3C']
POS_COLORS_MPL  = ['#E65100','#EF9A9A','#BF360C','#FFCCBC','#FF6D00']

def make_chart_image(title, metric, color_list, hdr_color, cd):
    months = cd["months"]
    years  = cd["years"]
    fig, ax = plt.subplots(figsize=(10, 3.5))
    for i, yr in enumerate(years):
        vals = cd["series"][str(yr)][metric]
        ax.plot(months, vals,
                color=color_list[i % len(color_list)],
                linewidth=2.5, marker='o', markersize=4,
                label=str(yr))
    ax.set_title(title, fontsize=12, fontweight='bold',
                 color='white', pad=8,
                 bbox=dict(facecolor=hdr_color, edgecolor='none',
                           boxstyle='round,pad=0.4'))
    ax.set_xticks(range(12))
    ax.set_xticklabels(months, fontsize=9)
    ax.yaxis.set_major_formatter(
        mticker.FuncFormatter(
            lambda x,_: f"{x/1e6:.1f}M" if abs(x)>=1e6
                        else f"{x/1e3:.1f}K" if abs(x)>=1e3
                        else f"{int(x):,}"))
    ax.tick_params(axis='y', labelsize=9)
    ax.grid(axis='y', color='#e0e0e0', linewidth=0.7)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.legend(fontsize=9, loc='upper right')
    fig.tight_layout(pad=1.2)
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=150, bbox_inches='tight')
    plt.close(fig)
    buf.seek(0)
    return buf

def xl_fill(hex_col):
    return PatternFill("solid", fgColor=hex_col)
thin  = Side(style="thin", color="AAAAAA")
bdr   = Border(left=thin,right=thin,top=thin,bottom=thin)
ctr   = Alignment(horizontal="center",vertical="center",wrap_text=True)
lft   = Alignment(horizontal="left",vertical="center",wrap_text=True)
rgt   = Alignment(horizontal="right",vertical="center")
wh_b  = Font(color="FFFFFF",bold=True,size=10)
bld   = Font(bold=True,size=10)
bld_i = Font(bold=True,italic=True,size=10)
nrm   = Font(size=10)

def xl_title(ws, text, hex_col, ncols):
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    c = ws["A1"]
    c.value = text
    c.fill  = xl_fill(hex_col)
    c.font  = Font(color="FFFFFF",bold=True,size=12)
    c.alignment = ctr
    ws.row_dimensions[1].height = 25

def xl_headers(ws, headers, widths, hex_col, row=2):
    for ci,(h,w) in enumerate(zip(headers,widths),1):
        c = ws.cell(row,ci,h)
        c.fill=xl_fill(hex_col); c.font=wh_b
        c.alignment=ctr; c.border=bdr
        ws.column_dimensions[get_column_letter(ci)].width=w
    ws.row_dimensions[row].height=30

def xl_data_row(ws, row, vals, hex_fill, num_cols=None):
    rf = xl_fill(hex_fill)
    for ci,v in enumerate(vals,1):
        c = ws.cell(row,ci)
        c.value = v
        c.fill=rf; c.border=bdr; c.font=nrm
        if ci==2: c.alignment=lft
        else: c.alignment=ctr
        if num_cols and ci in num_cols:
            c.number_format="#,##0"

def xl_tot_row(ws, row, vals, hex_fill, num_cols=None, italic=False):
    rf = xl_fill(hex_fill)
    ft = bld_i if italic else bld
    for ci,v in enumerate(vals,1):
        c = ws.cell(row,ci)
        c.value = v
        c.fill=rf; c.border=bdr; c.font=ft
        c.alignment=ctr
        if ci==1: c.alignment=rgt
        if num_cols and ci in num_cols:
            c.number_format="#,##0"

def build_excel(tbl_sel,e1,e2,e3,e4,e5,tot1,tot2,tot3,tot4,tot5):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    vn = st.session_state.vname_select
    yr = st.session_state.year_select
    sub = f"Vendor: {vn}  |  Year: {yr}  |  {datetime.now().strftime('%Y-%m-%d %H:%M')}"

    if "1" in tbl_sel:
        ws = wb.create_sheet("T1 - Top X Qty")
        xl_title(ws,"Top X Items Purchased - Qty  |  "+sub,"1F3864",4)
        xl_headers(ws,["Item No","Description","Total Qty","Rank"],[14,55,15,8],"1F3864")
        rf=[COL["r1a"],COL["r1b"]]
        for ri,r in enumerate(e1.to_dict("records"),3):
            xl_data_row(ws,ri,[r["ItemNo"],r["Desc"],r["TQty"],r["Rank"]],rf[ri%2],[3])
        xl_tot_row(ws,len(e1)+3,["Total","",tot1["TQty"],""],COL["tot"],[3])
        ws.freeze_panes="A3"

    if "2" in tbl_sel:
        ws = wb.create_sheet("T2 - Freq Purchased")
        xl_title(ws,"Top Frequently Purchased Items  |  "+sub,"005F73",6)
        xl_headers(ws,["Item No","Description","Total Qty","No of POs","Purch Factor","Rank"],
                   [14,50,13,11,13,8],"005F73")
        rf=[COL["r2a"],COL["r2b"]]
        for ri,r in enumerate(e2.to_dict("records"),3):
            xl_data_row(ws,ri,[r["ItemNo"],r["Desc"],r["TQty"],
                               r["NPOs"],round(r["PF"],4),r["Rank"]],rf[ri%2],[3,4])
        xl_tot_row(ws,len(e2)+3,["Total","",tot2["TQty"],tot2["NPOs"],"",""],COL["tot"],[3,4])
        ws.freeze_panes="A3"

    if "3" in tbl_sel:
        ws = wb.create_sheet("T3 - Price Analysis")
        xl_title(ws,"Analysis of Unit Price Changes  |  "+sub,"4A235A",10)
        xl_headers(ws,["Item No","Description","Total Qty","Total Cost","Avg Unit Price",
                        "Min Price","Max Price","Last Price","No of POs","Price Var"],
                   [14,50,12,14,14,12,12,12,10,12],"4A235A")
        rf=[COL["r3a"],COL["r3b"]]
        for ri,r in enumerate(e3.to_dict("records"),3):
            xl_data_row(ws,ri,[r["ItemNo"],r["Desc"],r["TQty"],r["TCost"],
                               round(r["AvgP"],2),round(r["MinP"],2),round(r["MaxP"],2),
                               round(r["LastP"],2),r["NPOs"],round(r["PVar"],2)],
                        rf[ri%2],[3,4])
        xl_tot_row(ws,len(e3)+3,["Total","",tot3["TQty"],tot3["TCost"],"","","","","",""],
                   COL["tot"],[3,4])
        ws.freeze_panes="A3"

    if "4" in tbl_sel:
        ws = wb.create_sheet("T4 - ABC Analysis")
        xl_title(ws,"ABC Spend Analysis  |  "+sub,"784212",6)
        xl_headers(ws,["Item No","Description","Total Cost","% of Spend","Cumulative %","ABC"],
                   [14,50,14,12,13,7],"784212")
        rf=[COL["r4a"],COL["r4b"]]
        for ri,r in enumerate(e4.to_dict("records"),3):
            xl_data_row(ws,ri,[r["ItemNo"],r["Desc"],r["TCost"],
                               round(r["PctS"],2),round(r["CumP"],2),r["ABC"]],
                        rf[ri%2],[3])
        xl_tot_row(ws,len(e4)+3,["Total","",tot4["TCost"],"","",""],COL["tot"],[3])
        ws.freeze_panes="A3"

    if "5" in tbl_sel:
        ws = wb.create_sheet("T5 - Monthly Trend")
        xl_title(ws,"Monthly Spend Trend  |  "+sub,"1A5276",5)
        xl_headers(ws,["Year / Month","No of POs","Total Qty","Total Cost","Avg Unit Price"],
                   [14,12,15,16,15],"1A5276")
        rf=[COL["r5a"],COL["r5b"]]
        ri=3
        for yr_val in sorted(e5["Year"].unique()):
            yr_rows=e5[e5["Year"]==yr_val].to_dict("records")
            yn=0;yq=0;yc=0
            for r in yr_rows:
                xl_data_row(ws,ri,[r["YearMonth"],r["NPOs"],r["TQty"],
                                   r["TCost"],round(r["AvgP"],2)],rf[ri%2],[3,4])
                yn+=r["NPOs"];yq+=r["TQty"];yc+=r["TCost"]; ri+=1
            xl_tot_row(ws,ri,[f"{yr_val} Total",yn,yq,yc,""],
                       COL["yrt"],[2,3,4],italic=True); ri+=1
        xl_tot_row(ws,ri,["Grand Total",tot5["NPOs"],tot5["TQty"],tot5["TCost"],""],
                   COL["tot"],[2,3,4])
        ws.freeze_panes="A3"

    if "C" in tbl_sel:
        months_hdr=["Jan","Feb","Mar","Apr","May","Jun",
                    "Jul","Aug","Sep","Oct","Nov","Dec"]
        cd=chart_data
        for sheet_title,metric,hex_col,color_list in [
            ("Chart1 - Cost YoY",   "cost","1F3864",COST_COLORS_MPL),
            ("Chart2 - Qty YoY",    "qty", "1B5E20",QTY_COLORS_MPL),
            ("Chart3 - POs YoY",    "npos","E65100",POS_COLORS_MPL),
        ]:
            ws = wb.create_sheet(sheet_title)
            ncols = 1+len(months_hdr)
            ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
            c=ws["A1"]
            c.value=sheet_title+" | "+sub
            c.fill=xl_fill(hex_col)
            c.font=Font(color="FFFFFF",bold=True,size=12)
            c.alignment=ctr
            ws.row_dimensions[1].height=25
            ws.cell(2,1,"Year").fill=xl_fill(hex_col)
            ws.cell(2,1).font=wh_b
            ws.cell(2,1).alignment=ctr
            ws.cell(2,1).border=bdr
            ws.column_dimensions["A"].width=10
            for ci,m in enumerate(months_hdr,2):
                c=ws.cell(2,ci,m)
                c.fill=xl_fill(hex_col);c.font=wh_b
                c.alignment=ctr;c.border=bdr
                ws.column_dimensions[get_column_letter(ci)].width=10
            rf=[COL["r5a"],COL["r5b"]]
            data_start_row=3
            for ri2,yr_val in enumerate(cd["years"],data_start_row):
                c=ws.cell(ri2,1,str(yr_val))
                c.fill=xl_fill(rf[ri2%2]);c.font=nrm;c.alignment=ctr;c.border=bdr
                for ci2,v in enumerate(cd["series"][str(yr_val)][metric],2):
                    cell=ws.cell(ri2,ci2,v)
                    cell.fill=xl_fill(rf[ri2%2])
                    cell.font=nrm;cell.alignment=ctr;cell.border=bdr
                    cell.number_format="#,##0"
            data_end_row=data_start_row+len(cd["years"])-1
            img_buf=make_chart_image(sheet_title,metric,color_list,"#"+hex_col,cd)
            img=openpyxl.drawing.image.Image(img_buf)
            img.width=900
            img.height=340
            ws.add_image(img,f"A{data_end_row+2}")
            ws.freeze_panes="A3"

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return bytes(buf.getvalue())

def hx(col):
    return colors.HexColor(f"#{col}" if not col.startswith("#") else col)

desc_style = ParagraphStyle("ds",fontSize=7.5,fontName="Helvetica",
    textColor=colors.black,leading=9,alignment=TA_LEFT)
cell_style = ParagraphStyle("cs",fontSize=7.5,fontName="Helvetica",
    textColor=colors.black,leading=9,alignment=TA_CENTER)

def ph(txt):
    return Paragraph(f'<font color="white"><b>{txt}</b></font>',
        ParagraphStyle("ph",fontSize=7.5,alignment=TA_CENTER,
                       fontName="Helvetica-Bold",textColor=colors.white,leading=9))
def PD(txt): return Paragraph(str(txt), desc_style)
def PC(txt): return Paragraph(str(txt), cell_style)

def pdf_sec(label, hex_col, W):
    t=Table([[Paragraph(f'<font color="white"><b>{label}</b></font>',
        ParagraphStyle("s",fontSize=10,alignment=TA_CENTER,
            fontName="Helvetica-Bold",textColor=colors.white))]],colWidths=[W])
    t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),hx(hex_col)),
        ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5)]))
    return t

def pdf_tbl(rows,col_w,hex_col,yr_rows=[],tot_rows=[],ra="D9E1F2",rb="EBF0FA"):
    t=Table(rows,colWidths=col_w,repeatRows=1)
    ts_=[
        ("BACKGROUND",(0,0),(-1,0),hx(hex_col)),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1),7.5),
        ("ALIGN",(0,0),(-1,-1),"CENTER"),
        ("ALIGN",(1,1),(1,-1),"LEFT"),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("GRID",(0,0),(-1,-1),0.3,hx("AAAAAA")),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[hx(ra),hx(rb)]),
        ("TOPPADDING",(0,0),(-1,-1),3),
        ("BOTTOMPADDING",(0,0),(-1,-1),3),
    ]
    for r in yr_rows:
        ts_+=[("BACKGROUND",(0,r),(-1,r),hx("AED6F1")),
              ("FONTNAME",(0,r),(-1,r),"Helvetica-BoldOblique"),
              ("LINEABOVE",(0,r),(-1,r),0.5,hx("7FB3D3"))]
    for r in tot_rows:
        ts_+=[("BACKGROUND",(0,r),(-1,r),hx("C0C0C0")),
              ("FONTNAME",(0,r),(-1,r),"Helvetica-Bold"),
              ("LINEABOVE",(0,r),(-1,r),1,colors.black)]
    t.setStyle(TableStyle(ts_))
    return t

def build_pdf(tbl_sel,e1,e2,e3,e4,e5,tot1,tot2,tot3,tot4,tot5):
    buf=io.BytesIO()
    doc=SimpleDocTemplate(buf,pagesize=landscape(A4),
        leftMargin=1.2*cm,rightMargin=1.2*cm,
        topMargin=1.2*cm,bottomMargin=1.2*cm)
    story=[]
    vn=st.session_state.vname_select
    yr=st.session_state.year_select
    W=25*cm
    fm=lambda v: f"{int(round(float(v or 0))):,}"
    ff=lambda v,d=2: f"{float(v or 0):.{d}f}"

    tit=Table([[Paragraph(
        f'<font color="white"><b>Cloud Energy Gas (CEG) — Vendor Purchase Analysis'
        f'  |  Vendor: {vn}  |  Year: {yr}  |  '
        f'{datetime.now().strftime("%Y-%m-%d %H:%M")}</b></font>',
        ParagraphStyle("t",fontSize=9,alignment=TA_CENTER,
            fontName="Helvetica-Bold",textColor=colors.white))]],colWidths=[W])
    tit.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),hx("1F3864")),
        ("TOPPADDING",(0,0),(-1,-1),7),("BOTTOMPADDING",(0,0),(-1,-1),7)]))
    story.append(tit); story.append(Spacer(1,0.3*cm))

    if "1" in tbl_sel:
        story.append(pdf_sec("Top X Items Purchased - Qty","1F3864",W))
        rows=[[ph("Item No"),ph("Description"),ph("Total Qty"),ph("Rank")]]
        for r in e1.to_dict("records"):
            rows.append([PC(r["ItemNo"]),PD(r["Desc"]),PC(fm(r["TQty"])),PC(str(r["Rank"]))])
        rows.append([PC(""),PC("Total"),PC(fm(tot1["TQty"])),PC("")])
        story.append(pdf_tbl(rows,[2.5*cm,14*cm,3*cm,2*cm],"1F3864",
            tot_rows=[len(rows)-1],ra=COL["r1a"],rb=COL["r1b"]))
        story.append(Spacer(1,0.4*cm))

    if "2" in tbl_sel:
        story.append(pdf_sec("Top Frequently Purchased Items","005F73",W))
        rows=[[ph("Item No"),ph("Description"),ph("Total Qty"),
               ph("No of POs"),ph("Purch Factor"),ph("Rank")]]
        for r in e2.to_dict("records"):
            rows.append([PC(r["ItemNo"]),PD(r["Desc"]),PC(fm(r["TQty"])),
                         PC(str(r["NPOs"])),PC(ff(r["PF"],4)),PC(str(r["Rank"]))])
        rows.append([PC(""),PC("Total"),PC(fm(tot2["TQty"])),
                     PC(fm(tot2["NPOs"])),PC(""),PC("")])
        story.append(pdf_tbl(rows,[2.2*cm,11*cm,2.5*cm,2.2*cm,2.8*cm,1.8*cm],"005F73",
            tot_rows=[len(rows)-1],ra=COL["r2a"],rb=COL["r2b"]))
        story.append(Spacer(1,0.4*cm))

    if "3" in tbl_sel:
        story.append(pdf_sec("Analysis of Unit Price Changes","4A235A",W))
        rows=[[ph("Item No"),ph("Description"),ph("Total Qty"),ph("Total Cost"),
               ph("Avg Unit Price"),ph("Min Price"),ph("Max Price"),
               ph("Last Price"),ph("No of POs"),ph("Price Var")]]
        for r in e3.to_dict("records"):
            rows.append([PC(r["ItemNo"]),PD(r["Desc"]),PC(fm(r["TQty"])),
                         PC(fm(r["TCost"])),PC(ff(r["AvgP"])),PC(ff(r["MinP"])),
                         PC(ff(r["MaxP"])),PC(ff(r["LastP"])),
                         PC(str(r["NPOs"])),PC(ff(r["PVar"]))])
        rows.append([PC(""),PC("Total"),PC(fm(tot3["TQty"])),PC(fm(tot3["TCost"])),
                     PC(""),PC(""),PC(""),PC(""),PC(""),PC("")])
        story.append(pdf_tbl(rows,
            [1.8*cm,7.5*cm,2.2*cm,2.2*cm,2*cm,2*cm,2*cm,2*cm,1.6*cm,1.8*cm],
            "4A235A",tot_rows=[len(rows)-1],ra=COL["r3a"],rb=COL["r3b"]))
        story.append(Spacer(1,0.4*cm))

    if "4" in tbl_sel:
        story.append(pdf_sec("ABC Spend Analysis","784212",W))
        rows=[[ph("Item No"),ph("Description"),ph("Total Cost"),
               ph("% of Spend"),ph("Cumulative %"),ph("ABC")]]
        for r in e4.to_dict("records"):
            rows.append([PC(r["ItemNo"]),PD(r["Desc"]),PC(fm(r["TCost"])),
                         PC(ff(r["PctS"])+"%"),PC(ff(r["CumP"])+"%"),PC(r["ABC"])])
        rows.append([PC(""),PC("Total"),PC(fm(tot4["TCost"])),PC(""),PC(""),PC("")])
        story.append(pdf_tbl(rows,[2.2*cm,12*cm,3*cm,2.5*cm,2.8*cm,1.5*cm],
            "784212",tot_rows=[len(rows)-1],ra=COL["r4a"],rb=COL["r4b"]))
        story.append(Spacer(1,0.4*cm))

    if "5" in tbl_sel:
        story.append(pdf_sec("Monthly Spend Trend","1A5276",W))
        rows=[[ph("Year / Month"),ph("No of POs"),ph("Total Qty"),
               ph("Total Cost"),ph("Avg Unit Price")]]
        yr_row_idx=[]
        for yr_val in sorted(e5["Year"].unique()):
            yr_rows_df=e5[e5["Year"]==yr_val].to_dict("records")
            for r in yr_rows_df:
                rows.append([PC(r["YearMonth"]),PC(str(r["NPOs"])),
                             PC(fm(r["TQty"])),PC(fm(r["TCost"])),PC(ff(r["AvgP"]))])
            yn=sum(r["NPOs"] for r in yr_rows_df)
            yq=sum(r["TQty"] for r in yr_rows_df)
            yc=sum(r["TCost"] for r in yr_rows_df)
            yr_row_idx.append(len(rows))
            rows.append([PC(f"{yr_val} Total"),PC(fm(yn)),PC(fm(yq)),PC(fm(yc)),PC("")])
        tot_idx=len(rows)
        rows.append([PC("Grand Total"),PC(fm(tot5["NPOs"])),
                     PC(fm(tot5["TQty"])),PC(fm(tot5["TCost"])),PC("")])
        story.append(pdf_tbl(rows,[3*cm,2.5*cm,3.5*cm,3.5*cm,3*cm],
            "1A5276",yr_rows=yr_row_idx,tot_rows=[tot_idx],
            ra=COL["r5a"],rb=COL["r5b"]))
        story.append(Spacer(1,0.4*cm))

    if "C" in tbl_sel:
        cd=chart_data
        months_hdr=["Jan","Feb","Mar","Apr","May","Jun",
                    "Jul","Aug","Sep","Oct","Nov","Dec"]
        col_w_ch=[2*cm]+[1.8*cm]*12
        for label,metric,hcol,color_list in [
            ("Monthly Cost - Year on Year",      "cost","1F3864",COST_COLORS_MPL),
            ("Monthly Qty - Year on Year",       "qty", "1B5E20",QTY_COLORS_MPL),
            ("Monthly No of POs - Year on Year", "npos","E65100",POS_COLORS_MPL),
        ]:
            story.append(pdf_sec(label,hcol,W))
            rows=[[ph("Year")]+[ph(m) for m in months_hdr]]
            for yr_val in cd["years"]:
                vals=[PC(str(yr_val))]+[PC(fm(v)) for v in cd["series"][str(yr_val)][metric]]
                rows.append(vals)
            story.append(pdf_tbl(rows,col_w_ch,hcol,
                ra=COL["r5a"],rb=COL["r5b"]))
            story.append(Spacer(1,0.3*cm))
            img_buf=make_chart_image(label,metric,color_list,"#"+hcol,cd)
            rl_img=RLImage(img_buf,width=W,height=8*cm)
            story.append(rl_img)
            story.append(Spacer(1,0.5*cm))

    doc.build(story)
    buf.seek(0)
    return bytes(buf.getvalue())

if do_export:
    if not export_tables:
        st.warning("Please select at least one table.")
    elif not export_fmt:
        st.warning("Please select at least one format.")
    else:
        tbl_sel=[]
        if any("Table 1" in t for t in export_tables): tbl_sel.append("1")
        if any("Table 2" in t for t in export_tables): tbl_sel.append("2")
        if any("Table 3" in t for t in export_tables): tbl_sel.append("3")
        if any("Table 4" in t for t in export_tables): tbl_sel.append("4")
        if any("Table 5" in t for t in export_tables): tbl_sel.append("5")
        if any("Charts" in t for t in export_tables):  tbl_sel.append("C")
        e1,e2,e3,e4,e5,tot1,tot2,tot3,tot4,tot5 = get_export_data()
        vn=st.session_state.vname_select
        yr=st.session_state.year_select
        ts=datetime.now().strftime("%Y%m%d_%H%M")
        if "Excel" in export_fmt:
            st.session_state.xl_data  = bytes(build_excel(
                tbl_sel,e1,e2,e3,e4,e5,tot1,tot2,tot3,tot4,tot5))
            st.session_state.xl_fname = f"CEG_{vn}_{yr}_{ts}.xlsx"
            st.session_state.xl_ready = True
        if "PDF" in export_fmt:
            st.session_state.pdf_data  = bytes(build_pdf(
                tbl_sel,e1,e2,e3,e4,e5,tot1,tot2,tot3,tot4,tot5))
            st.session_state.pdf_fname = f"CEG_{vn}_{yr}_{ts}.pdf"
            st.session_state.pdf_ready = True

if st.session_state.xl_ready and st.session_state.xl_data:
    xl_b64 = b64lib.b64encode(st.session_state.xl_data).decode()
    xl_fname = st.session_state.xl_fname
    st.markdown(f"""
        <a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{xl_b64}"
           download="{xl_fname}"
           style="display:inline-block;padding:8px 20px;background:#1F3864;color:white;
                  text-decoration:none;border-radius:5px;font-size:13px;font-weight:bold;
                  margin:4px 4px 4px 0;">
            &#128190; Download Excel &mdash; {xl_fname}
        </a>
    """, unsafe_allow_html=True)
    if st.button("✖ Clear Excel download", key="clr_xl"):
        st.session_state.xl_ready = False
        st.session_state.xl_data  = None
        st.rerun()

if st.session_state.pdf_ready and st.session_state.pdf_data:
    pdf_b64 = b64lib.b64encode(st.session_state.pdf_data).decode()
    pdf_fname = st.session_state.pdf_fname
    st.markdown(f"""
        <a href="data:application/pdf;base64,{pdf_b64}"
           download="{pdf_fname}"
           style="display:inline-block;padding:8px 20px;background:#784212;color:white;
                  text-decoration:none;border-radius:5px;font-size:13px;font-weight:bold;
                  margin:4px 4px 4px 0;">
            &#128190; Download PDF &mdash; {pdf_fname}
        </a>
    """, unsafe_allow_html=True)
    if st.button("✖ Clear PDF download", key="clr_pdf"):
        st.session_state.pdf_ready = False
        st.session_state.pdf_data  = None
        st.rerun()

html_content = """
<style>
*{box-sizing:border-box;margin:0;padding:0;font-family:Arial,sans-serif}
html,body{background:#f8f9fa;color:#000;height:100%;overflow:hidden}
.page{display:flex;flex-direction:column;height:100vh;padding:5px}
.top-section{flex:0 0 auto;overflow:hidden;background:#f8f9fa}
.top-scroll{overflow-x:scroll;overflow-y:hidden;height:12px;width:100%}
.top-scroll-inner{height:1px}
.hdr-container{overflow:hidden;width:100%}
.hdr-inner{display:flex;flex-direction:row;gap:20px;min-width:max-content}
.tblock{flex:0 0 auto;display:flex;flex-direction:column;overflow:hidden}
.sec-hdr{width:100%;height:42px;display:flex;align-items:center;
    justify-content:center;border-radius:5px 5px 0 0;
    color:white;font-weight:bold;font-size:20px;}
.tf{display:flex;align-items:center;gap:6px;height:38px;background:#f8f9fa}
.st{font-size:11px;color:gray;height:14px;line-height:14px;
    background:#f8f9fa;margin-bottom:1px}
.fh-tbl{border-collapse:collapse;font-size:12px;table-layout:fixed;width:100%}
.fh-tbl th{padding:4px 5px;text-align:center;color:white !important;
    font-size:11px;border:1px solid rgba(255,255,255,0.3);
    height:44px;white-space:normal;word-wrap:break-word;
    line-height:1.2;vertical-align:middle;}
.data-section{flex:1 1 auto;overflow-x:scroll;overflow-y:auto;width:100%}
.data-inner{display:flex;flex-direction:row;gap:20px;
    min-width:max-content;padding-bottom:40px;padding-right:30px;
    align-items:flex-start}
.dtb{flex:0 0 auto;overflow:hidden}
.dt{border-collapse:collapse;font-size:12px;table-layout:fixed;
    background:#fff;width:100%}
.dt td{padding:5px 7px;text-align:center;font-size:12px;
    color:#000;border:1px solid #e0e0e0}
.tw{word-wrap:break-word;white-space:normal}
.tn{white-space:nowrap}
.tot td{font-weight:bold;background:#e8e8e8 !important;
    color:#000 !important;border-top:2px solid #666 !important}
.yrtot td{font-weight:bold;background:#d0e4f7 !important;
    color:#000 !important;border-top:1px solid #aaa !important;font-style:italic;}
.fl{background:#FFD700;padding:0 8px;border-radius:4px;font-weight:bold;
    font-size:11px;height:28px;line-height:28px;white-space:nowrap;
    color:#000;display:inline-block}
.nw{display:flex;align-items:center;border:1px solid #d0d3da;
    border-radius:4px;overflow:hidden;height:28px;background:#fff}
.nw button{background:#f0f2f6;border:none;width:26px;height:28px;
    cursor:pointer;font-size:15px;color:#1F3864;font-weight:bold;
    transition:background 0.2s}
.nw button:hover{background:#FF6B00;color:white}
.nw input{width:42px;border:none;text-align:center;font-size:12px;
    color:#000;background:#fff;outline:none}
.ab{background:#f0f2f6;border:1px solid #d0d3da;border-radius:4px;
    padding:0 8px;height:28px;cursor:pointer;font-size:11px;
    font-weight:bold;color:#000;transition:background 0.2s}
.ab:hover{background:#FF6B00;color:white;border-color:#FF6B00}
.srch{border:1px solid #d0d3da;border-radius:4px;padding:0 8px;
    height:28px;font-size:11px;color:#000;background:#fff;
    outline:none;width:100px}
.abcsel{border:1px solid #d0d3da;border-radius:4px;padding:0 4px;
    height:28px;font-size:11px;color:#000;background:#fff;
    outline:none;cursor:pointer;width:70px}
.ch-card{background:#fff;border:1px solid #ddd;border-radius:5px;
    overflow:hidden;width:380px;margin-bottom:10px}
.ch-card:last-child{margin-bottom:0}
.ch-hdr{padding:9px 10px;color:#fff;font-weight:bold;font-size:14px;text-align:center}
.ch-body{padding:10px 10px 6px 10px}
.ch-legend{display:flex;gap:14px;justify-content:center;margin-top:6px;flex-wrap:wrap}
.ch-leg-item{display:flex;align-items:center;gap:5px;font-size:11px;color:#333}
.ch-leg-line{width:20px;height:3px;border-radius:2px;display:inline-block}
</style>

<div class="page">
  <div class="top-section">
    <div class="top-scroll" id="topScroll">
      <div class="top-scroll-inner" id="topScrollInner"></div>
    </div>
    <div class="hdr-container">
    <div class="hdr-inner" id="hdrInner">

      <div class="tblock" style="width:485px">
        <div class="sec-hdr" style="background:#1F3864">&#128230; Top X items purchased - qty</div>
        <div class="tf">
          <div class="fl">Filter Top X Rank</div>
          <div class="nw">
            <button onclick="adj('t1',-1)">&#8722;</button>
            <input type="number" id="n-t1" value="10" min="1"
              onchange="sa.t1=false;ra();sessionStorage.setItem('f_t1_n',this.value);sessionStorage.setItem('f_t1_all','false')">
            <button onclick="adj('t1',1)">+</button>
          </div>
          <button class="ab" onclick="sa.t1=true;ra();sessionStorage.setItem('f_t1_all','true')">All</button>
        </div>
        <div class="st" id="s1">Top 10</div>
        <table class="fh-tbl">
          <colgroup><col style="width:80px"><col style="width:270px">
          <col style="width:80px"><col style="width:55px"></colgroup>
          <tr><th style="background:#1F3864">Item No</th>
          <th style="background:#1F3864">Description</th>
          <th style="background:#1F3864">Total Qty</th>
          <th style="background:#1F3864">Rank</th></tr>
        </table>
      </div>

      <div class="tblock" style="width:600px">
        <div class="sec-hdr" style="background:#005F73">&#128203; Top frequently purchased items</div>
        <div class="tf">
          <div class="fl">Filter Freq PO Rank</div>
          <div class="nw">
            <button onclick="adj('t2',-1)">&#8722;</button>
            <input type="number" id="n-t2" value="10" min="1"
              onchange="sa.t2=false;ra();sessionStorage.setItem('f_t2_n',this.value);sessionStorage.setItem('f_t2_all','false')">
            <button onclick="adj('t2',1)">+</button>
          </div>
          <button class="ab" onclick="sa.t2=true;ra();sessionStorage.setItem('f_t2_all','true')">All</button>
        </div>
        <div class="st" id="s2">Top 10</div>
        <table class="fh-tbl">
          <colgroup><col style="width:80px"><col style="width:270px">
          <col style="width:80px"><col style="width:55px">
          <col style="width:65px"><col style="width:50px"></colgroup>
          <tr><th style="background:#005F73">Item No</th>
          <th style="background:#005F73">Description</th>
          <th style="background:#005F73">Total Qty</th>
          <th style="background:#005F73">No of POs</th>
          <th style="background:#005F73">Purch Factor</th>
          <th style="background:#005F73">Rank</th></tr>
        </table>
      </div>

      <div class="tblock" style="width:925px">
        <div class="sec-hdr" style="background:#4A235A">&#128200; Analysis of unit price changes</div>
        <div class="tf">
          <div class="fl">Filter Item Rank</div>
          <div class="nw">
            <button onclick="adj('t3',-1)">&#8722;</button>
            <input type="number" id="n-t3" value="10" min="1"
              onchange="sa.t3=false;ra();sessionStorage.setItem('f_t3_n',this.value);sessionStorage.setItem('f_t3_all','false')">
            <button onclick="adj('t3',1)">+</button>
          </div>
          <button class="ab" onclick="sa.t3=true;ra();sessionStorage.setItem('f_t3_all','true')">All</button>
        </div>
        <div class="st" id="s3">Top 10</div>
        <table class="fh-tbl">
          <colgroup><col style="width:80px"><col style="width:270px">
          <col style="width:80px"><col style="width:80px">
          <col style="width:75px"><col style="width:75px">
          <col style="width:75px"><col style="width:75px">
          <col style="width:45px"><col style="width:70px"></colgroup>
          <tr><th style="background:#4A235A">Item No</th>
          <th style="background:#4A235A">Description</th>
          <th style="background:#4A235A">Total Qty</th>
          <th style="background:#4A235A">Total Cost</th>
          <th style="background:#4A235A">Avg Unit Price</th>
          <th style="background:#4A235A">Min Price</th>
          <th style="background:#4A235A">Max Price</th>
          <th style="background:#4A235A">Last Price</th>
          <th style="background:#4A235A">No of POs</th>
          <th style="background:#4A235A">Price Var</th></tr>
        </table>
      </div>

      <div class="tblock" style="width:620px">
        <div class="sec-hdr" style="background:#784212">&#128181; ABC Spend Analysis</div>
        <div class="tf">
          <div class="fl">Item Filter</div>
          <input class="srch" type="text" id="t4-search" placeholder="Search..."
            oninput="ra4();sessionStorage.setItem('f_t4_srch',this.value)">
          <div class="fl">ABC</div>
          <select class="abcsel" id="t4-abc"
            onchange="ra4();sessionStorage.setItem('f_t4_abc',this.value)">
            <option value="ALL">ALL</option>
            <option value="AB">A &amp; B</option>
            <option value="A">A</option>
            <option value="B">B</option>
            <option value="C">C</option>
          </select>
        </div>
        <div class="st" id="s4">Showing All</div>
        <table class="fh-tbl">
          <colgroup><col style="width:80px"><col style="width:270px">
          <col style="width:80px"><col style="width:68px">
          <col style="width:72px"><col style="width:50px"></colgroup>
          <tr><th style="background:#784212">Item No</th>
          <th style="background:#784212">Description</th>
          <th style="background:#784212">Total Cost</th>
          <th style="background:#784212">% of Spend</th>
          <th style="background:#784212">Cumulative %</th>
          <th style="background:#784212">ABC</th></tr>
        </table>
      </div>

      <div class="tblock" style="width:385px">
        <div class="sec-hdr" style="background:#1A5276">&#128197; Monthly Spend Trend</div>
        <div class="tf"></div>
        <div class="st">&nbsp;</div>
        <table class="fh-tbl">
          <colgroup><col style="width:85px"><col style="width:65px">
          <col style="width:70px"><col style="width:80px">
          <col style="width:85px"></colgroup>
          <tr><th style="background:#1A5276">Year / Month</th>
          <th style="background:#1A5276">No of POs</th>
          <th style="background:#1A5276">Total Qty</th>
          <th style="background:#1A5276">Total Cost</th>
          <th style="background:#1A5276">Avg Unit Price</th></tr>
        </table>
      </div>

      <div style="width:3px;background:#ccc;border-radius:2px;
          align-self:stretch;flex-shrink:0;margin:0 5px"></div>

      <div class="tblock" style="width:380px">
        <div class="sec-hdr" style="background:#1F3864">&#128200; Year on Year Comparison Charts</div>
        <div class="tf"></div>
        <div class="st">&nbsp;</div>
        <div style="height:44px"></div>
      </div>

    </div>
    </div>
  </div>

  <div class="data-section" id="dataSection">
  <div class="data-inner">
    <div class="dtb" style="width:485px"><div id="tb1"></div></div>
    <div class="dtb" style="width:600px"><div id="tb2"></div></div>
    <div class="dtb" style="width:925px"><div id="tb3"></div></div>
    <div class="dtb" style="width:620px"><div id="tb4"></div></div>
    <div class="dtb" style="width:385px"><div id="tb5"></div></div>

    <div style="width:3px;background:#ccc;border-radius:2px;
        align-self:stretch;margin:0 5px;flex-shrink:0"></div>

    <div style="width:380px;flex-shrink:0;display:flex;flex-direction:column;gap:10px">
      <div class="ch-card">
        <div class="ch-hdr" style="background:#1F3864">&#128200; Monthly Cost — Year on Year</div>
        <div class="ch-body">
          <canvas id="chartCost" height="150"></canvas>
          <div class="ch-legend" id="legCost"></div>
        </div>
      </div>
      <div class="ch-card">
        <div class="ch-hdr" style="background:#1B5E20">&#128200; Monthly Qty — Year on Year</div>
        <div class="ch-body">
          <canvas id="chartQty" height="150"></canvas>
          <div class="ch-legend" id="legQty"></div>
        </div>
      </div>
      <div class="ch-card">
        <div class="ch-hdr" style="background:#E65100">&#128200; Monthly No of POs — Year on Year</div>
        <div class="ch-body">
          <canvas id="chartPOs" height="150"></canvas>
          <div class="ch-legend" id="legPOs"></div>
        </div>
      </div>
    </div>

  </div>
  </div>
</div>

<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<script>
const D = DATA_PLACEHOLDER;
let sa={t1:false,t2:false,t3:false};

const COST_COLORS=['#1F3864','#85B7EB','#4A235A','#005F73','#1A5276'];
const QTY_COLORS =['#1B5E20','#66BB6A','#2E7D32','#A5D6A7','#388E3C'];
const POS_COLORS =['#E65100','#EF9A9A','#BF360C','#FFCCBC','#FF6D00'];

let chartCostInst=null,chartQtyInst=null,chartPOsInst=null;

function fmtK(v){
    if(Math.abs(v)>=1000000)return(v/1000000).toFixed(1)+'M';
    if(Math.abs(v)>=1000)return(v/1000).toFixed(1)+'K';
    return Math.round(v).toLocaleString();
}

function buildLegend(el,years,colorArr){
    el.innerHTML='';
    years.forEach((yr,i)=>{
        const col=colorArr[i%colorArr.length];
        el.innerHTML+=`<div class="ch-leg-item">
            <span class="ch-leg-line" style="background:${col}"></span>
            <span>${yr}</span></div>`;
    });
}

function buildChart(canvasId,years,seriesData,colorArr,labelFn){
    const ctx=document.getElementById(canvasId).getContext('2d');
    const datasets=years.map((yr,i)=>({
        label:String(yr),
        data:seriesData[String(yr)],
        borderColor:colorArr[i%colorArr.length],
        backgroundColor:'transparent',
        borderWidth:2.5,
        pointRadius:3,
        pointHoverRadius:5,
        tension:0.3,
    }));
    return new Chart(ctx,{
        type:'line',
        data:{labels:D.charts.months,datasets},
        options:{
            responsive:true,
            plugins:{
                legend:{display:false},
                tooltip:{callbacks:{label:ctx=>`${ctx.dataset.label}: ${labelFn(ctx.raw)}`}}
            },
            scales:{
                x:{ticks:{font:{size:10}},grid:{color:'rgba(0,0,0,0.05)'}},
                y:{ticks:{font:{size:10},callback:v=>labelFn(v)},
                   grid:{color:'rgba(0,0,0,0.05)'}}
            }
        }
    });
}

function initCharts(){
    const cd=D.charts;
    const years=cd.years;
    const costData={},qtyData={},posData={};
    years.forEach(yr=>{
        costData[String(yr)]=cd.series[String(yr)].cost;
        qtyData[String(yr)] =cd.series[String(yr)].qty;
        posData[String(yr)] =cd.series[String(yr)].npos;
    });
    if(chartCostInst)chartCostInst.destroy();
    if(chartQtyInst) chartQtyInst.destroy();
    if(chartPOsInst) chartPOsInst.destroy();
    chartCostInst=buildChart('chartCost',years,costData,COST_COLORS,fmtK);
    chartQtyInst =buildChart('chartQty', years,qtyData, QTY_COLORS, fmtK);
    chartPOsInst =buildChart('chartPOs', years,posData, POS_COLORS,
        v=>Math.round(v).toLocaleString());
    buildLegend(document.getElementById('legCost'),years,COST_COLORS);
    buildLegend(document.getElementById('legQty'), years,QTY_COLORS);
    buildLegend(document.getElementById('legPOs'), years,POS_COLORS);
}

window.addEventListener('load',function(){
    const topScroll=document.getElementById('topScroll');
    const hdrInner=document.getElementById('hdrInner');
    const dataSection=document.getElementById('dataSection');
    function updateWidth(){
        document.getElementById('topScrollInner').style.width=hdrInner.scrollWidth+'px';
        topScroll.style.width=dataSection.clientWidth+'px';
    }
    updateWidth();
    window.addEventListener('resize',updateWidth);
    let syncing=false;
    topScroll.addEventListener('scroll',function(){
        if(syncing)return;syncing=true;
        hdrInner.style.marginLeft='-'+topScroll.scrollLeft+'px';
        dataSection.scrollLeft=topScroll.scrollLeft;
        syncing=false;
    });
    dataSection.addEventListener('scroll',function(){
        if(syncing)return;syncing=true;
        topScroll.scrollLeft=dataSection.scrollLeft;
        hdrInner.style.marginLeft='-'+dataSection.scrollLeft+'px';
        syncing=false;
    });
    setTimeout(initCharts,300);
});

function n(t){return parseInt(document.getElementById('n-'+t).value)||10;}
function adj(t,d){
    const el=document.getElementById('n-'+t);
    const v=parseInt(el.value)||10;
    if(v+d>=1){
        el.value=v+d;sa[t]=false;ra();
        sessionStorage.setItem('f_'+t+'_n',v+d);
        sessionStorage.setItem('f_'+t+'_all','false');
    }
}
function fmt(v){return Math.round(Number(v||0)).toLocaleString();}
function ff(v,d=2){return Number(v||0).toFixed(d);}
function ffc(v,d=2){return Number(v||0).toLocaleString('en-US',{minimumFractionDigits:d,maximumFractionDigits:d});}

function cg1(){return '<colgroup><col style="width:80px"><col style="width:270px"><col style="width:80px"><col style="width:55px"></colgroup>';}
function cg2(){return '<colgroup><col style="width:80px"><col style="width:270px"><col style="width:80px"><col style="width:55px"><col style="width:65px"><col style="width:50px"></colgroup>';}
function cg3(){return '<colgroup><col style="width:80px"><col style="width:270px"><col style="width:80px"><col style="width:80px"><col style="width:75px"><col style="width:75px"><col style="width:75px"><col style="width:75px"><col style="width:45px"><col style="width:70px"></colgroup>';}
function cg4(){return '<colgroup><col style="width:80px"><col style="width:270px"><col style="width:80px"><col style="width:68px"><col style="width:72px"><col style="width:50px"></colgroup>';}
function cg5(){return '<colgroup><col style="width:85px"><col style="width:65px"><col style="width:70px"><col style="width:80px"><col style="width:85px"></colgroup>';}

function t1(data){
    const rows=sa.t1?data:data.slice(0,n('t1'));
    document.getElementById('s1').textContent=sa.t1?'Showing All':'Top '+n('t1');
    let h='<table class="dt">'+cg1()+'<tbody>';
    rows.forEach((r,i)=>{
        const bg=i%2===0?'#D9E1F2':'#EBF0FA';
        h+='<tr style="background:'+bg+'"><td class="tn">'+r.ItemNo+'</td><td class="tw">'+r.Desc+'</td><td class="tn">'+fmt(r.TQty)+'</td><td class="tn">'+r.Rank+'</td></tr>';
    });
    const tot=rows.reduce((s,r)=>s+Number(r.TQty),0);
    h+='<tr class="tot"><td colspan="2" style="text-align:right">Total</td><td class="tn">'+fmt(tot)+'</td><td></td></tr>';
    document.getElementById('tb1').innerHTML=h+'</tbody></table>';
}
function t2(data){
    const rows=sa.t2?data:data.slice(0,n('t2'));
    document.getElementById('s2').textContent=sa.t2?'Showing All':'Top '+n('t2');
    let h='<table class="dt">'+cg2()+'<tbody>';
    rows.forEach((r,i)=>{
        const bg=i%2===0?'#D4EFEA':'#E8F8F5';
        h+='<tr style="background:'+bg+'"><td class="tn">'+r.ItemNo+'</td><td class="tw">'+r.Desc+'</td><td class="tn">'+fmt(r.TQty)+'</td><td class="tn">'+r.NPOs+'</td><td class="tn">'+ff(r.PF,4)+'</td><td class="tn">'+r.Rank+'</td></tr>';
    });
    const tq=rows.reduce((s,r)=>s+Number(r.TQty),0);
    const tn=rows.reduce((s,r)=>s+Number(r.NPOs),0);
    h+='<tr class="tot"><td colspan="2" style="text-align:right">Total</td><td class="tn">'+fmt(tq)+'</td><td class="tn">'+fmt(tn)+'</td><td></td><td></td></tr>';
    document.getElementById('tb2').innerHTML=h+'</tbody></table>';
}
function t3(data){
    const rows=sa.t3?data:data.slice(0,n('t3'));
    document.getElementById('s3').textContent=sa.t3?'Showing All':'Top '+n('t3');
    let h='<table class="dt">'+cg3()+'<tbody>';
    rows.forEach((r,i)=>{
        const bg=i%2===0?'#E8DAEF':'#F5EEF8';
        h+='<tr style="background:'+bg+'">'
            +'<td class="tn">'+r.ItemNo+'</td>'
            +'<td class="tw">'+r.Desc+'</td>'
            +'<td class="tn">'+fmt(r.TQty)+'</td>'
            +'<td class="tn">'+fmt(r.TCost)+'</td>'
            +'<td class="tn">'+ffc(r.AvgP)+'</td>'
            +'<td class="tn">'+ffc(r.MinP)+'</td>'
            +'<td class="tn">'+ffc(r.MaxP)+'</td>'
            +'<td class="tn">'+ffc(r.LastP)+'</td>'
            +'<td class="tn">'+r.NPOs+'</td>'
            +'<td class="tn">'+ffc(r.PVar)+'</td>'
            +'</tr>';
    });
    const tq=rows.reduce((s,r)=>s+Number(r.TQty),0);
    const tc=rows.reduce((s,r)=>s+Number(r.TCost),0);
    h+='<tr class="tot"><td colspan="2" style="text-align:right">Total</td><td class="tn">'+fmt(tq)+'</td><td class="tn">'+fmt(tc)+'</td><td colspan="6"></td></tr>';
    document.getElementById('tb3').innerHTML=h+'</tbody></table>';
}
function ra4(){
    const srch=document.getElementById('t4-search').value.toLowerCase();
    const abc=document.getElementById('t4-abc').value;
    let rows=D.t4.filter(r=>{
        const mi=srch===''||r.ItemNo.toLowerCase().includes(srch)||r.Desc.toLowerCase().includes(srch);
        const ma=abc==='ALL'||(abc==='AB'&&(r.ABC==='A'||r.ABC==='B'))||r.ABC===abc;
        return mi&&ma;
    });
    document.getElementById('s4').textContent='Showing '+rows.length+' items';
    let h='<table class="dt">'+cg4()+'<tbody>';
    rows.forEach((r,i)=>{
        const bg=i%2===0?'#FAD7A0':'#FDEBD0';
        h+='<tr style="background:'+bg+'"><td class="tn">'+r.ItemNo+'</td><td class="tw">'+r.Desc+'</td><td class="tn">'+fmt(r.TCost)+'</td><td class="tn">'+ff(r.PctS)+'%</td><td class="tn">'+ff(r.CumP)+'%</td><td class="tn">'+r.ABC+'</td></tr>';
    });
    const filtTot=rows.reduce((s,r)=>s+Number(r.TCost),0);
    h+='<tr class="tot"><td colspan="2" style="text-align:right">Total</td><td class="tn">'+fmt(filtTot)+'</td><td colspan="3"></td></tr>';
    document.getElementById('tb4').innerHTML=h+'</tbody></table>';
}
function t5(data){
    let h='<table class="dt">'+cg5()+'<tbody>';
    const years=[...new Set(data.map(r=>r.Year))].sort();
    years.forEach(yr=>{
        const yrows=data.filter(r=>r.Year===yr);
        let yN=0,yQ=0,yC=0;
        yrows.forEach((r,i)=>{
            const bg=i%2===0?'#D6EAF8':'#EBF5FB';
            yN+=Number(r.NPOs);yQ+=Number(r.TQty);yC+=Number(r.TCost);
            h+='<tr style="background:'+bg+'"><td class="tn">'+r.YearMonth+'</td><td class="tn">'+r.NPOs+'</td><td class="tn">'+fmt(r.TQty)+'</td><td class="tn">'+fmt(r.TCost)+'</td><td class="tn">'+ff(r.AvgP)+'</td></tr>';
        });
        h+='<tr class="yrtot"><td class="tn" style="text-align:right">'+yr+' Total</td><td class="tn">'+fmt(yN)+'</td><td class="tn">'+fmt(yQ)+'</td><td class="tn">'+fmt(yC)+'</td><td></td></tr>';
    });
    const gN=data.reduce((s,r)=>s+Number(r.NPOs),0);
    const gQ=data.reduce((s,r)=>s+Number(r.TQty),0);
    const gC=data.reduce((s,r)=>s+Number(r.TCost),0);
    h+='<tr class="tot"><td style="text-align:right">Grand Total</td><td class="tn">'+fmt(gN)+'</td><td class="tn">'+fmt(gQ)+'</td><td class="tn">'+fmt(gC)+'</td><td></td></tr>';
    document.getElementById('tb5').innerHTML=h+'</tbody></table>';
}
function ra(){t1(D.t1);t2(D.t2);t3(D.t3);ra4();t5(D.t5);}
ra();
</script>
"""

html_content = html_content.replace("DATA_PLACEHOLDER", data_json)
st.components.v1.html(html_content, height=1800, scrolling=True)
